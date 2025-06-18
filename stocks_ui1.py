import streamlit as st
from io import BytesIO
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import yfinance as yf
import logging
import mpmath as mp
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

mp.dps = 100

logging.basicConfig(
    level=logging.DEBUG, 
    format="%(asctime)s | %(levelname)s | %(filename)s:%(lineno)d | %(message)s",
    handlers=[logging.StreamHandler()]
)

# Getting data from Yahoo finance
def get_data(stock_name, start_date, end_date):
    data: pd.DataFrame = yf.download(stock_name, start=start_date, end=end_date, auto_adjust=False)
    closing_prices = data['Close']
    s_obs = closing_prices.to_numpy()
    return s_obs, data

def determine_v_n(Sn, Sn_1):
    v_n = (Sn - Sn_1) / 1
    if abs(v_n) < 1e-12:
        return 1e-12 
    return v_n

def determine_alpha_n(Sn_minus_4, Sn_minus_3, Sn_minus_2, Sn_minus_1):
    AA = (Sn_minus_2 - 2 * Sn_minus_3 + Sn_minus_4)
    BB = (Sn_minus_1 - Sn_minus_2)
    CC = (Sn_minus_1 - 2 * Sn_minus_2 + Sn_minus_3)
    DD = (Sn_minus_2 - Sn_minus_3)
    alpha_numerator = (AA * BB) - (CC * DD)
    alpha_denominator = DD * BB * (DD - BB) 
    if abs(alpha_denominator) < 1e-12:
        return 1e-12  
    return (alpha_numerator / alpha_denominator)

def determine_beta_n(Sn_minus_3, Sn_minus_2, Sn_minus_1, alpha_n):
    CC = (Sn_minus_1 - 2 * Sn_minus_2 + Sn_minus_3)
    BB = (Sn_minus_1 - Sn_minus_2)
    if abs(BB) < 1e-12:
        return 1e-12 
    return (CC - (alpha_n * (BB**2))) / (BB * 1)

def determine_h_n(v_1, alpha_n, beta_n):
    if abs(alpha_n) < 1e-12:
        alpha_n = 1e-12
    if abs(v_1) < 1e-12:
        v_1 = 1e-12
    try:
        h_n = abs((v_1 + (beta_n / alpha_n) / v_1))
        return h_n
    except (ZeroDivisionError) as e:
        logging.warning(f"Error in determine_h_n: {e}. Using fallback value.")
        return 1.0

def determine_s_n(s1, alpha, beta, h, condition_1, s_n, v_n, v_1):
    logging.debug(f"determine_s_n called with: s1={s1}, alpha={alpha}, beta={beta}, h={h}, condition_1={condition_1}, s_n={s_n}, v_n={v_n}, v_1={v_1}")
    if abs(alpha) < 1e-12:
        alpha = 1e-12
    if abs(beta) < 1e-12:
        beta = 1e-12
    condition_2 = v_n > v_1
    condition_3 = s_n > s1
    try:
        if condition_1 > 0 and condition_2 and condition_3:
            s_n = s1 - (1/alpha) * mp.log(mp.fabs((mp.exp(beta) - h) / (1 - h)))
        if condition_1 > 0 and condition_2 and not condition_3:
            s_n = s1 + mp.fabs(1/alpha) * (mp.fabs(beta)/beta) * mp.log(mp.fabs((mp.exp(beta) - h) / (1 - h)))
        if condition_1 < 0 and condition_2 and condition_3:
            s_n = s1 - (1/alpha) * mp.log(mp.fabs((mp.exp(beta) + h) / (1 + h)))
        if condition_1 < 0 and condition_2 and not condition_3:
            s_n = s1 - mp.fabs(1/alpha) * (mp.fabs(beta)/beta) * mp.log(mp.fabs((mp.exp(beta) + h) / (1 + h)))
        if condition_1 > 0 and not condition_2 and condition_3:
            s_n = s1 - (1/alpha) * (beta/mp.fabs(beta)) * mp.log(mp.fabs((mp.exp(beta) - h) / (1 - h)))
        if condition_1 > 0 and not condition_2 and not condition_3:
            s_n = s1 - mp.fabs(1/alpha) * mp.log(mp.fabs((mp.exp(-mp.fabs(beta)) - h) / (1 - h)))
        if condition_1 < 0 and not condition_2 and condition_3:
            s_n = s1 + (1/alpha) * (beta/mp.fabs(beta)) * mp.log(mp.fabs(mp.exp(-mp.fabs(beta)) + h) / (1 + h))
        if condition_1 < 0 and not condition_2 and not condition_3:
            s_n = s1 + mp.fabs(1/alpha) * mp.log(mp.fabs(mp.exp(-mp.fabs(beta)) + h) / (1 + h))
    except (ZeroDivisionError) as e:
        logging.error(f'Error in determine_s_n: {e}. Using fallback value.')
        s_n = s1
    logging.debug(f'determine_s_n result: s_n={s_n}')
    return s_n

def determine_MAPE_list(actual: list, predicted: list) -> list:
    logging.debug(f'actual: {actual}, len {len(actual)}')
    logging.debug(f"predicted: {predicted}, len {len(predicted)}")
    min_len = min(len(actual), len(predicted))
    actual = actual[:min_len]
    predicted = predicted[:min_len]
    num_of_cases = len(actual)
    sum_of_percentage_error = 0
    mape_list = []
    for i in range(num_of_cases):
        if actual[i] == 0:
            continue
        abs_error = mp.fabs(actual[i] - predicted[i])
        percentage_error = abs_error / actual[i]
        sum_of_percentage_error += percentage_error
        MAPE = sum_of_percentage_error / (i + 1) * 100
        mape_list.append(float(MAPE))
    return mape_list

def fitting(closing_prices, stock_symbol):
    logging.debug(f'fitting called with closing_prices={closing_prices}, stock_symbol={stock_symbol}')
    Fitting_S_n_list = []
    v_list = []
    first_run = True
    if len(closing_prices) < 4:
        st.error("Not enough data to perform fitting. At least 4 data points are required.")
        return [], []
    for i in range(3):
        Fitting_S_n_list.append(float(closing_prices[i]))
    for i in range(3, len(closing_prices)):
        S_minus_1 = closing_prices[i - 3]
        S_0 = closing_prices[i - 2]
        S_1 = closing_prices[i - 1]
        S_2 = closing_prices[i]
        v_0 = determine_v_n(S_0, S_minus_1)
        v_1 = determine_v_n(S_1, S_0)
        v_2 = determine_v_n(S_2, S_1)
        if first_run:
            v_list.append(v_0)
            v_list.append(v_1)
            first_run = False
        v_list.append(v_2)
        try:
            alpha_n = determine_alpha_n(S_minus_1, S_0, S_1, S_2)
            beta_n = determine_beta_n(S_minus_1, S_1, S_2, alpha_n)
            h_n = determine_h_n(v_0, alpha_n, beta_n)
            condition_1 = (v_2 + (beta_n / alpha_n)) * v_2
            S_n = determine_s_n(S_minus_1, alpha_n, beta_n, h_n, condition_1, S_2, v_2, v_0)
        except (ZeroDivisionError) as e:
            logging.warning(f"Error in calculation at index {i}: {e}. Using fallback.")
            S_n = S_2
        Fitting_S_n_list.append(float(S_n))
        logging.debug(f'Appended S_n={S_n} to Fitting_S_n_list')
    return Fitting_S_n_list, v_list

def forecasting(Fitting_S_n_list, start_date, end_date, stock_symbol):
    if len(Fitting_S_n_list) < 4:
        st.error("Not enough fitting data to perform forecasting.")
        return [], []
    fitting_S_last = Fitting_S_n_list[-4:].copy()
    try:
        closing_prices_full = get_data(stock_symbol, start_date, end_date)[0]
        closing_prices_full = [price.item() for price in closing_prices_full]
        closing_prices_full = filter_prices_duplicates(closing_prices_full)
    except Exception as e:
        st.error(f"Error getting forecast data: {e}")
        return [], []
    forecast_days = len(closing_prices_full) - len(Fitting_S_n_list)
    if forecast_days <= 0:
        st.warning("Not enough data to perform forecasting.")
        return [], closing_prices_full[len(Fitting_S_n_list):]
    for i in range(3, forecast_days + 3):
        if i >= len(fitting_S_last):
            break
        S_minus_1 = fitting_S_last[i - 3]
        S_0 = fitting_S_last[i - 2]
        S_1 = fitting_S_last[i - 1]
        S_2 = fitting_S_last[i]
        v_0 = determine_v_n(S_0, S_minus_1)
        v_2 = determine_v_n(S_2, S_1)
        try:
            alpha_n = determine_alpha_n(S_minus_1, S_0, S_1, S_2)
            beta_n = determine_beta_n(S_0, S_1, S_2, alpha_n)
            h_n = determine_h_n(v_0, alpha_n, beta_n)
            condition_1 = (v_2 + (beta_n / alpha_n)) * v_2
            S_n = determine_s_n(S_minus_1, alpha_n, beta_n, h_n, condition_1, S_2, v_2, v_0)
        except (ZeroDivisionError) as e:
            logging.warning(f"Error in forecast at step {i}: {e}. Using previous value.")
            S_n = S_2
        fitting_S_last.append(float(S_n))
    forecast_S_list = fitting_S_last[3:] 
    closing_forecast = closing_prices_full[len(Fitting_S_n_list)-1:]
    return forecast_S_list, closing_forecast

def filter_prices_duplicates(closing_prices):
    if not closing_prices:
        return []
    filtered_prices = [closing_prices[0]]
    for i in range(1, len(closing_prices)):
        if closing_prices[i] != closing_prices[i-1]:
            filtered_prices.append(closing_prices[i])
    return filtered_prices

def create_excel_download(stock_symbol, start_date, end_date, forecast_end_date, 
                         closing_prices, Fitting_S_n_list, S_forecast, closing_forecast):
    try:
        fitting_data = yf.download(stock_symbol, start=start_date, end=end_date, auto_adjust=False)
        fitting_dates = fitting_data.index.tolist()
        if len(fitting_dates) > len(closing_prices):
            fitting_dates = fitting_dates[:len(closing_prices)]
        fitting_df = pd.DataFrame({
            'Date': fitting_dates,
            'Actual_Price': closing_prices,
            'Fitted_Price': Fitting_S_n_list[:len(closing_prices)] if len(Fitting_S_n_list) >= len(closing_prices) else Fitting_S_n_list + [None] * (len(closing_prices) - len(Fitting_S_n_list)),
            'Type': 'Fitting'
        })
        if S_forecast and closing_forecast:
            forecast_data = yf.download(stock_symbol, start=end_date, end=forecast_end_date, auto_adjust=False)
            forecast_dates = forecast_data.index.tolist()
            if forecast_dates and fitting_dates and forecast_dates[0] <= fitting_dates[-1]:
                forecast_dates = forecast_dates[1:]
            min_forecast_len = min(len(forecast_dates), len(closing_forecast), len(S_forecast))
            forecast_dates = forecast_dates[:min_forecast_len]
            closing_forecast = closing_forecast[:min_forecast_len]
            S_forecast = S_forecast[:min_forecast_len]
            forecast_df = pd.DataFrame({
                'Date': forecast_dates,
                'Actual_Price': closing_forecast,
                'Fitted_Price': None,
                'Forecast_Price': S_forecast,
                'Type': 'Forecast'
            })
            fitting_df['Forecast_Price'] = None
            combined_df = pd.concat([fitting_df, forecast_df], ignore_index=True)
        else:
            fitting_df['Forecast_Price'] = None
            combined_df = fitting_df
        combined_df = combined_df[['Date', 'Actual_Price', 'Fitted_Price', 'Forecast_Price', 'Type']]
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{stock_symbol}_Analysis"
        ws['A1'] = f"Stock Analysis Report - {stock_symbol}"
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Fitting Period: {start_date} to {end_date}"
        ws['A4'] = f"Forecast Period: {end_date} to {forecast_end_date}"
        ws['A5'] = ""
        title_font = Font(size=14, bold=True)
        ws['A1'].font = title_font
        headers = ['Date', 'Actual Price', 'Fitted Price', 'Forecast Price', 'Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for row_idx, (_, row) in enumerate(combined_df.iterrows(), 7):
            if pd.notna(row['Date']):
                if hasattr(row['Date'], 'strftime'):
                    ws.cell(row=row_idx, column=1, value=row['Date'].strftime('%Y-%m-%d'))
                else:
                    ws.cell(row=row_idx, column=1, value=str(row['Date']))
            else:
                ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value=row['Actual_Price'])
            ws.cell(row=row_idx, column=3, value=row['Fitted_Price'])
            ws.cell(row=row_idx, column=4, value=row['Forecast_Price'])
            ws.cell(row=row_idx, column=5, value=row['Type'])
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        logging.error(f"Error in create_excel_download: {e}")

# Streamlit UI - Main Page Layout
st.title("📈 Stock Price Fitting and Forecasting Web")

st.markdown("---")

# Input Parameter
st.subheader("📋 Input Parameters")

col1, col2, col3, col4 = st.columns(4)

with col1:
    st.markdown("**Stock Symbol**")
    stock_symbol = st.text_input("", value="BBCA.JK", key="stock_input", label_visibility="collapsed")

with col2:
    st.markdown("**Fitting Start Date**")
    start_date = st.date_input("", value=datetime(2024, 1, 1), key="start_date_input", label_visibility="collapsed")

with col3:
    st.markdown("**Fitting Period (Days)**")
    training_days = st.number_input("", min_value=1, max_value=365, value=120, step=1, key="training_days", label_visibility="collapsed")

with col4:
    st.markdown("**Forecast Period (Days)**")
    forecast_days = st.number_input("", min_value=1, max_value=365, value=60, step=1, key="forecast_days", label_visibility="collapsed")

def clear_all():
    st.session_state.stock_input = "BBCA.JK"
    st.session_state.start_date_input = datetime(2024, 1, 1)
    st.session_state.training_days = 120
    st.session_state.forecast_days = 60
    st.session_state.use_custom_end = False
    st.session_state.custom_end = datetime(2024, 4, 30)
    st.session_state.use_custom_forecast_end = False
    st.session_state.custom_forecast_end = datetime(2024, 6, 29)

st.button("🗑️ Clear All", on_click=clear_all, use_container_width=True)

end_date = start_date + timedelta(days=training_days)
forecast_end_date = end_date + timedelta(days=forecast_days)

with st.expander("⚙️ Advanced Options"):
    col_adv1, col_adv2 = st.columns(2)
    with col_adv1:
        use_custom_end = st.checkbox("Custom Fitting End Date", value=False, key="use_custom_end")
        if use_custom_end:
            st.markdown("**Custom Fitting End Date**")
            custom_end_date = st.date_input("", value=datetime(2024, 4, 30), key="custom_end", label_visibility="collapsed")
            end_date = custom_end_date
    with col_adv2:
        use_custom_forecast_end = st.checkbox("Custom Forecast End Date", value=False, key="use_custom_forecast_end")
        if use_custom_forecast_end:
            st.markdown("**Custom Forecast End Date**")
            custom_forecast_end = st.date_input("", value=datetime(2024, 6, 29), key="custom_forecast_end", label_visibility="collapsed")
            forecast_end_date = custom_forecast_end

col_summary1, col_summary2 = st.columns(2)

with col_summary1:
    st.markdown(f"""
    **📊 Fitting Period Details:**
    - **Fitting Period:** {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}
    - **Fitting Duration:** {training_days} days
    """)

with col_summary2:
    st.markdown(f"""
    **🔮 Forecast Period Details:**
    - **Forecast Period:** {end_date.strftime('%d/%m/%Y')} - {forecast_end_date.strftime('%d/%m/%Y')}
    - **Forecast Duration:** {forecast_days} days
    """)

st.markdown("---")

if start_date >= end_date:
    st.error("Start date must be earlier than end date!")
elif end_date >= forecast_end_date:
    st.error("End date must be earlier than forecast end date!")

run_forecast = st.button("🔗 Submit Data", use_container_width=True, type="primary")

if run_forecast:
    try:
        with st.spinner("Fetching and processing data..."):
            _, raw_data = get_data(stock_symbol, start_date, end_date)
            if raw_data.empty:
                st.error("No data retrieved from Yahoo Finance. Please check the stock symbol or date range.")
                st.stop()

            st.subheader(f"📋 Raw Data from Yahoo Finance ({stock_symbol})")
            st.markdown(f"Data retrieved for the period: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}")
            raw_data_display = raw_data.reset_index()
            raw_data_display['Date'] = raw_data_display['Date'].dt.strftime('%Y-%m-%d')
            st.dataframe(
                raw_data_display,
                use_container_width=True,
                column_config={
                    "Date": st.column_config.DateColumn("Date"),
                    "Open": st.column_config.NumberColumn("Open", format="%.2f"),
                    "High": st.column_config.NumberColumn("High", format="%.2f"),
                    "Low": st.column_config.NumberColumn("Low", format="%.2f"),
                    "Close": st.column_config.NumberColumn("Close", format="%.2f"),
                    "Adj Close": st.column_config.NumberColumn("Adj Close", format="%.2f"),
                    "Volume": st.column_config.NumberColumn("Volume", format="%d")
                }
            )
            st.markdown("This table shows the raw data retrieved from Yahoo Finance. Use it to verify if the data is complete (e.g., no missing dates or values) before fitting and forecasting.")

            closing_prices, _ = get_data(stock_symbol, start_date, end_date)
            closing_prices = [price.item() for price in closing_prices]
            closing_prices = filter_prices_duplicates(closing_prices)
            if len(closing_prices) < 4:
                st.error("Not enough data to perform forecasting. At least 4 data points are required.")
                st.stop()
            Fitting_S_n_list, v_list = fitting(closing_prices, stock_symbol)
            if not Fitting_S_n_list:
                st.error("Failed to fit data.")
                st.stop()
            mape_fit = determine_MAPE_list(closing_prices, Fitting_S_n_list)
            S_forecast, closing_forecast = forecasting(
                Fitting_S_n_list,
                start_date.strftime("%Y-%m-%d"),
                forecast_end_date.strftime("%Y-%m-%d"),
                stock_symbol
            )
            if S_forecast and closing_forecast:
                mape_forecast = determine_MAPE_list(closing_forecast, S_forecast)
            else:
                mape_forecast = []

        st.success("Done!")
        if Fitting_S_n_list:
            st.subheader("📊 Statistic Details")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Number of Fitting Data Points", len(closing_prices))
            with col2:
                if mape_fit:
                    st.metric("MAPE Fitting", f"{np.mean(mape_fit):.2f}%")
            with col3:
                if mape_forecast:
                    st.metric("MAPE Forecast", f"{np.mean(mape_forecast):.2f}%")
            with col4:
                st.metric("Forecast Period", f"{(forecast_end_date - end_date).days} days")

            # Fitting vs Actual Chart and Table
            st.subheader(f"📊 Fitting vs Actual Chart ({stock_symbol})")
            fig_fit, ax_fit = plt.subplots(figsize=(10, 6))
            ax_fit.plot(closing_prices, label="Actual", color='black', linewidth=2)
            ax_fit.plot(Fitting_S_n_list, label="Fitted", color='blue', linewidth=2)
            ax_fit.set_title(f"Stock Price Fitting Data ({stock_symbol})")
            ax_fit.set_xlabel("Day")
            ax_fit.set_ylabel("Price")
            ax_fit.legend()
            ax_fit.grid(True, alpha=0.3)
            st.pyplot(fig_fit)

            # Table for Fitting vs Actual
            fitting_data = yf.download(stock_symbol, start=start_date, end=end_date, auto_adjust=False)
            fitting_dates = fitting_data.index.tolist()[:len(closing_prices)]
            fitting_df = pd.DataFrame({
                'Date': [d.strftime('%Y-%m-%d') for d in fitting_dates],
                'Actual Price': closing_prices,
                'Fitted Price': Fitting_S_n_list[:len(closing_prices)]
            })
            st.subheader(f"📋 Fitting vs Actual Data Table ({stock_symbol})")
            st.dataframe(
                fitting_df,
                use_container_width=True,
                column_config={
                    "Date": st.column_config.DateColumn("Date"),
                    "Actual Price": st.column_config.NumberColumn("Actual Price", format="%.2f"),
                    "Fitted Price": st.column_config.NumberColumn("Fitted Price", format="%.2f")
                }
            )

            # Fitting + Forecast vs Actual Chart and Table
            if S_forecast and closing_forecast:
                st.subheader(f"📈 Fitting + Forecast vs Actual Chart ({stock_symbol})")
                fig_forecast, ax_forecast = plt.subplots(figsize=(12, 6))
                all_actual = closing_prices + closing_forecast
                ax_forecast.plot(all_actual, label="Actual", color='black', linewidth=2)
                ax_forecast.plot(range(len(Fitting_S_n_list)), Fitting_S_n_list, label="Fitted", color='blue', linewidth=2)
                forecast_start_idx = len(Fitting_S_n_list)
                forecast_end_idx = forecast_start_idx + len(S_forecast)
                ax_forecast.plot(range(forecast_start_idx, forecast_end_idx), S_forecast, 
                               label="Forecast", color='orange', linewidth=2)
                ax_forecast.axvline(x=len(closing_prices), color='red', linestyle='--', 
                                  label='Forecast Start', alpha=0.7)
                ax_forecast.set_title(f"Stock Price Fitting and Forecast ({stock_symbol})")
                ax_forecast.set_xlabel("Day")
                ax_forecast.set_ylabel("Price")
                ax_forecast.legend()
                ax_forecast.grid(True, alpha=0.3)
                st.pyplot(fig_forecast)

                # Table for Fitting + Forecast vs Actual
                forecast_data = yf.download(stock_symbol, start=end_date, end=forecast_end_date, auto_adjust=False)
                forecast_dates = forecast_data.index.tolist()
                if forecast_dates and fitting_dates and forecast_dates[0] <= fitting_dates[-1]:
                    forecast_dates = forecast_dates[1:]
                min_forecast_len = min(len(forecast_dates), len(closing_forecast), len(S_forecast))
                forecast_dates = forecast_dates[:min_forecast_len]
                closing_forecast = closing_forecast[:min_forecast_len]
                S_forecast = S_forecast[:min_forecast_len]
                combined_df = pd.DataFrame({
                    'Date': [d.strftime('%Y-%m-%d') for d in fitting_dates + forecast_dates],
                    'Actual Price': closing_prices + closing_forecast,
                    'Fitted Price': Fitting_S_n_list + [None] * len(forecast_dates),
                    'Forecast Price': [None] * len(fitting_dates) + S_forecast
                })
                st.subheader(f"📋 Fitting + Forecast vs Actual Data Table ({stock_symbol})")
                st.dataframe(
                    combined_df,
                    use_container_width=True,
                    column_config={
                        "Date": st.column_config.DateColumn("Date"),
                        "Actual Price": st.column_config.NumberColumn("Actual Price", format="%.2f"),
                        "Fitted Price": st.column_config.NumberColumn("Fitted Price", format="%.2f"),
                        "Forecast Price": st.column_config.NumberColumn("Forecast Price", format="%.2f")
                    }
                )

            # MAPE Fitting Chart and Table
            if mape_fit:
                st.subheader(f"📉 MAPE Fitting Results - Average: {np.mean(mape_fit):.2f}%")
                fig_mape_fit, ax_mape_fit = plt.subplots(figsize=(10, 6))
                ax_mape_fit.plot(mape_fit, color='purple', label='MAPE Fitting (%)', linewidth=2)
                ax_mape_fit.set_title(f"MAPE Chart During Fitting ({stock_symbol})")
                ax_mape_fit.set_xlabel("Day")
                ax_mape_fit.set_ylabel("MAPE (%)")
                ax_mape_fit.legend()
                ax_mape_fit.grid(True, alpha=0.3)
                st.pyplot(fig_mape_fit)

                # Table for MAPE Fitting
                mape_fit_df = pd.DataFrame({
                    'Date': [d.strftime('%Y-%m-%d') for d in fitting_dates[:len(mape_fit)]],
                    'MAPE (%)': mape_fit
                })
                st.subheader(f"📋 MAPE Fitting Data Table ({stock_symbol})")
                st.dataframe(
                    mape_fit_df,
                    use_container_width=True,
                    column_config={
                        "Date": st.column_config.DateColumn("Date"),
                        "MAPE (%)": st.column_config.NumberColumn("MAPE (%)", format="%.2f")
                    }
                )

            # MAPE Forecast Chart and Table
            if mape_forecast:
                st.subheader(f"📉 MAPE Forecast Results - Average: {np.mean(mape_forecast):.2f}%")
                fig_mape_forecast, ax_mape_forecast = plt.subplots(figsize=(10, 6))
                ax_mape_forecast.plot(mape_forecast, color='orange', label='MAPE Forecast (%)', linewidth=2)
                ax_mape_forecast.set_title(f"MAPE Chart During Forecasting ({stock_symbol})")
                ax_mape_forecast.set_xlabel("Day")
                ax_mape_forecast.set_ylabel("MAPE (%)")
                ax_mape_forecast.legend()
                ax_mape_forecast.grid(True, alpha=0.3)
                st.pyplot(fig_mape_forecast)

                # Table for MAPE Forecast
                mape_forecast_df = pd.DataFrame({
                    'Date': [d.strftime('%Y-%m-%d') for d in forecast_dates[:len(mape_forecast)]],
                    'MAPE (%)': mape_forecast
                })
                st.subheader(f"📋 MAPE Forecast Data Table ({stock_symbol})")
                st.dataframe(
                    mape_forecast_df,
                    use_container_width=True,
                    column_config={
                        "Date": st.column_config.DateColumn("Date"),
                        "MAPE (%)": st.column_config.NumberColumn("MAPE (%)", format="%.2f")
                    }
                )

            st.subheader("💾 Download Data")
            try:
                excel_data = create_excel_download(
                    stock_symbol=stock_symbol,
                    start_date=start_date,
                    end_date=end_date,
                    forecast_end_date=forecast_end_date,
                    closing_prices=closing_prices,
                    Fitting_S_n_list=Fitting_S_n_list,
                    S_forecast=S_forecast if S_forecast else [],
                    closing_forecast=closing_forecast if closing_forecast else []
                )
                filename = f"{stock_symbol}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download complete analysis data in Excel format"
                )
                st.info(f"📊 This file contains data from {start_date} to {forecast_end_date}")
            except Exception as e:
                st.error(f"Error creating Excel file: {str(e)}")
                logging.error(f"Excel creation error: {e}")
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
        logging.error(f"Main execution error: {e}")
        st.info("Please try with different parameters or check the data connection.")